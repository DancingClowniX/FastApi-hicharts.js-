
from fastapi.responses import FileResponse
from fastapi import FastAPI, HTTPException
import uvicorn
from fastapi.responses import JSONResponse
from fastapi import APIRouter, Request
from fastapi.templating import Jinja2Templates
from datetime import datetime
from typing import List, Union
from pydantic import BaseModel
from typing import Optional,List,Dict
import sqlite3
import os

import openpyxl
from fastapi.staticfiles import StaticFiles
app = FastAPI()

templates = Jinja2Templates(directory="templates")

app.mount('/static', StaticFiles(directory='static'), name='static')


@app.get("/")
async def get_html(request:Request):
    return templates.TemplateResponse('home.html',{"request":request})

@app.get("/data")
async def get_json(request:Request):
    #
    #
    # def export_to_sqlite():
    #     '''Экспорт данных из xlsx в sqlite'''
    #
    #     # 1. Создание и подключение к базе
    #
    #     # Получаем текущую папку проекта
    #     prj_dir = os.path.abspath(os.path.curdir)
    #
    #     a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    #
    #     # Имя базы
    #     base_name = 'auto.sqlite3'
    #
    #     # метод sqlite3.connect автоматически создаст базу, если ее нет
    #     connect = sqlite3.connect(prj_dir + '/' + base_name)
    #     # курсор - это специальный объект, который делает запросы и получает результаты запросов
    #     cursor = connect.cursor()
    #
    #     # создание таблицы если ее не существует
    #     cursor.execute('CREATE TABLE IF NOT EXISTS table_data (bp, id int, status, plan_date date,prognoz_date date, filial,check_plan, check_fact,plan_month int,plan_year int,min_date_fact_month int,min_date_fact_year int)')
    #
    #     # 2. Работа c xlsx файлом
    #
    #     # Читаем файл и лист1 книги excel
    #     file_to_read = openpyxl.load_workbook('bd_file.xlsx', data_only=True)
    #     sheet = file_to_read['Лист1']
    #
    #     # Цикл по строкам начиная со второй (в первой заголовки)
    #
    #     for row in range(2, sheet.max_row + 1):
    #         # Объявление списка
    #         data = []
    #         # Цикл по столбцам от 1 до 4 ( 5 не включая)
    #         for col in range(1, 13):
    #             # value содержит значение ячейки с координатами row col
    #             value = sheet.cell(row, col).value
    #             # Список который мы потом будем добавлять
    #             data.append(value)
    #
    #         # 3. Запись в базу и закрытие соединения
    #
    #         # Вставка данных в поля таблицы
    #         cursor.execute("INSERT INTO table_data VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3],data[4],data[5],data[6],data[7],data[8],data[9],data[10],data[11]))
    #
    #     # сохраняем изменения
    #     connect.commit()
    #     # закрытие соединения
    #     connect.close()
    # export_to_sqlite()
    def get_data():
        connect = sqlite3.connect(database='auto.sqlite3')
        con = connect.cursor()
        con.execute("""SELECT 
    filial,
    SUM(CASE WHEN plan_month = '1' THEN dt ELSE 0 END) AS "1",
    SUM(CASE WHEN plan_month = '2' THEN dt ELSE 0 END) AS "2",
    SUM(CASE WHEN plan_month = '3' THEN dt ELSE 0 END) AS "3",
    SUM(CASE WHEN plan_month = '4' THEN dt ELSE 0 END) AS "4",
    SUM(CASE WHEN plan_month = '5' THEN dt ELSE 0 END) AS "5",
    SUM(CASE WHEN plan_month = '6' THEN dt ELSE 0 END) AS "6",
    SUM(CASE WHEN plan_month = '7' THEN dt ELSE 0 END) AS "7",
    SUM(CASE WHEN plan_month = '8' THEN dt ELSE 0 END) AS "8",
    SUM(CASE WHEN plan_month = '9' THEN dt ELSE 0 END) AS "9",
    SUM(CASE WHEN plan_month = '10' THEN dt ELSE 0 END) AS "10",
    SUM(CASE WHEN plan_month = '11' THEN dt ELSE 0 END) AS "11",
    SUM(CASE WHEN plan_month = '12' THEN dt ELSE 0 END) AS "12"
FROM (
    SELECT 
        filial, 
        COUNT(CASE WHEN check_fact = 'Да' THEN 1 END) AS dt,
        plan_month
    FROM 
        table_data 
    WHERE 
        plan_year = 2024 
    GROUP BY 
        filial, plan_month
) AS subquery
GROUP BY 
    filial""")
        return con.fetchall()

    data = get_data()

    # Преобразуем данные в формат JSON
    response_data = []
    for row in data:
        response_data.append({
            "filial": row[0],
            "1": row[1],
            "2": row[2],
            "3": row[3],
            "4": row[4],
            "5": row[5],
            "6": row[6],
            "7": row[7],
            "8": row[8],
            "9": row[9],
            "10": row[10],
            "11": row[11],
            "12": row[12],
        })
    return JSONResponse(content=response_data)

